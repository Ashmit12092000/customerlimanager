from decimal import Decimal
from datetime import datetime, date
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from models import Transaction, Customer
import math

def calculate_interest(principal, annual_rate, days):
    """Calculate simple interest for given principal, rate and days"""
    if not principal or not annual_rate or not days:
        return Decimal('0')
    
    # Ensure all values are Decimal
    principal = Decimal(str(principal))
    annual_rate = Decimal(str(annual_rate))
    days = Decimal(str(days))
    
    # Simple interest formula: (P * R * T) / 100
    # Where T is in years (days/365)
    interest = (principal * annual_rate * days) / (Decimal('100') * Decimal('365'))
    return interest.quantize(Decimal('0.01'))

def calculate_compound_interest(principal, annual_rate, days, frequency):
    """Calculate compound interest for given principal, rate, days and frequency"""
    if not principal or not annual_rate or not days or not frequency:
        return Decimal('0')
    
    # Ensure all values are Decimal
    principal = Decimal(str(principal))
    annual_rate = Decimal(str(annual_rate))
    days = Decimal(str(days))
    
    # Convert frequency to compounding periods per year
    if frequency == 'monthly':
        n = 12
    elif frequency == 'quarterly':
        n = 4
    elif frequency == 'yearly':
        n = 1
    else:
        n = 1
    
    # Convert days to years
    t = days / Decimal('365')
    
    # Compound interest formula: P * (1 + r/n)^(nt) - P
    r = annual_rate / Decimal('100')
    n = Decimal(str(n))
    
    # Using decimal arithmetic for precision
    amount = principal * ((Decimal('1') + r/n) ** (n * t))
    compound_interest = amount - principal
    
    return compound_interest.quantize(Decimal('0.01'))

def export_to_excel(customer, transactions):
    """Export customer data and transactions to Excel"""
    output = io.BytesIO()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Customer Report - {customer.icl_no}"
    
    # Define styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Customer Information Header
    ws['A1'] = 'Customer Master Report'
    ws['A1'].font = title_font
    ws.merge_cells('A1:K1')
    
    # Customer Details
    row = 3
    ws[f'A{row}'] = 'ICL No:'
    ws[f'B{row}'] = customer.icl_no
    ws[f'A{row}'].font = header_font
    
    row += 1
    ws[f'A{row}'] = 'Customer Name:'
    ws[f'B{row}'] = customer.name
    ws[f'A{row}'].font = header_font
    
    row += 1
    ws[f'A{row}'] = 'Address:'
    ws[f'B{row}'] = customer.address
    ws[f'A{row}'].font = header_font
    
    row += 1
    ws[f'A{row}'] = 'Contact Details:'
    ws[f'B{row}'] = customer.contact_details
    ws[f'A{row}'].font = header_font
    
    row += 1
    ws[f'A{row}'] = 'Annual Rate:'
    ws[f'B{row}'] = f"{customer.annual_rate}%"
    ws[f'A{row}'].font = header_font
    
    row += 1
    ws[f'A{row}'] = 'Interest Type:'
    ws[f'B{row}'] = customer.interest_type.title()
    ws[f'A{row}'].font = header_font
    
    row += 1
    ws[f'A{row}'] = 'TDS Applicable:'
    ws[f'B{row}'] = 'Yes' if customer.tds_applicable else 'No'
    ws[f'A{row}'].font = header_font
    
    row += 1
    ws[f'A{row}'] = 'Current Balance:'
    ws[f'B{row}'] = customer.get_current_balance()
    ws[f'A{row}'].font = header_font
    
    # Transaction Headers
    row += 3
    headers = ['Date', 'Amount Paid', 'Amount Repaid', 'Balance', 'From', 'To', 
               'No of Days', 'Int Rate', 'Int Amount', 'TDS', 'Net Amount']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Transaction Data
    total_int_amount = Decimal('0')
    total_tds = Decimal('0')
    total_net = Decimal('0')
    total_days = 0
    
    for transaction in transactions:
        row += 1
        data = [
            transaction.date.strftime('%d-%m-%Y') if transaction.date else '',
            transaction.amount_paid or '',
            transaction.amount_repaid or '',
            transaction.balance or '',
            transaction.period_from.strftime('%d-%m-%Y') if transaction.period_from else '',
            transaction.period_to.strftime('%d-%m-%Y') if transaction.period_to else '',
            transaction.no_of_days or '',
            f"{transaction.int_rate}%" if transaction.int_rate else '',
            transaction.int_amount or '',
            transaction.tds_amount or '',
            transaction.net_amount or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col > 4:  # Align numbers to right
                cell.alignment = Alignment(horizontal='right')
        
        # Add to totals (with proper decimal handling)
        if transaction.int_amount is not None:
            total_int_amount += Decimal(str(transaction.int_amount))
        if transaction.tds_amount is not None:
            total_tds += Decimal(str(transaction.tds_amount))
        if transaction.net_amount is not None:
            total_net += Decimal(str(transaction.net_amount))
        if transaction.no_of_days:
            total_days += transaction.no_of_days
    
    # Totals row
    row += 1
    ws[f'A{row}'] = 'Total'
    ws[f'A{row}'].font = header_font
    ws[f'G{row}'] = total_days
    ws[f'I{row}'] = total_int_amount
    ws[f'J{row}'] = total_tds
    ws[f'K{row}'] = total_net
    
    # Apply borders to totals
    for col in range(1, 12):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).font = header_font
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

def get_period_report(start_date, end_date):
    """Generate period-based report for all customers"""
    output = io.BytesIO()
    
    # Get all transactions in the period
    transactions = Transaction.query.filter(
        Transaction.date >= start_date,
        Transaction.date <= end_date
    ).order_by(Transaction.date).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Period Report {start_date} to {end_date}"
    
    # Define styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws['A1'] = f'Period Report: {start_date.strftime("%d-%m-%Y")} to {end_date.strftime("%d-%m-%Y")}'
    ws['A1'].font = title_font
    ws.merge_cells('A1:L1')
    
    # Headers
    row = 3
    headers = ['Customer ICL', 'Customer Name', 'Date', 'Amount Paid', 'Amount Repaid', 
               'Balance', 'From', 'To', 'No of Days', 'Int Rate', 'Int Amount', 'TDS', 'Net Amount']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    total_int_amount = Decimal('0')
    total_tds = Decimal('0')
    total_net = Decimal('0')
    
    for transaction in transactions:
        row += 1
        data = [
            transaction.customer.icl_no,
            transaction.customer.name,
            transaction.date.strftime('%d-%m-%Y') if transaction.date else '',
            transaction.amount_paid or '',
            transaction.amount_repaid or '',
            transaction.balance or '',
            transaction.period_from.strftime('%d-%m-%Y') if transaction.period_from else '',
            transaction.period_to.strftime('%d-%m-%Y') if transaction.period_to else '',
            transaction.no_of_days or '',
            f"{transaction.int_rate}%" if transaction.int_rate else '',
            transaction.int_amount or '',
            transaction.tds_amount or '',
            transaction.net_amount or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col > 6:  # Align numbers to right
                cell.alignment = Alignment(horizontal='right')
        
        # Add to totals (with proper decimal handling)
        if transaction.int_amount is not None:
            total_int_amount += Decimal(str(transaction.int_amount))
        if transaction.tds_amount is not None:
            total_tds += Decimal(str(transaction.tds_amount))
        if transaction.net_amount is not None:
            total_net += Decimal(str(transaction.net_amount))
    
    # Totals row
    row += 1
    ws[f'A{row}'] = 'Total'
    ws[f'A{row}'].font = header_font
    ws[f'K{row}'] = total_int_amount
    ws[f'L{row}'] = total_tds
    ws[f'M{row}'] = total_net
    
    # Apply borders to totals
    for col in range(1, 14):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).font = header_font
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output
